import json
import pymysql
from .import models
import numpy as np
from django.shortcuts import render,redirect,HttpResponse, get_object_or_404
from django.http import HttpResponseNotAllowed, HttpResponseRedirect
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from openpyxl import load_workbook
from .forms import CoordinateForm
from django.http import JsonResponse
from django.db.models import Max
import torch
import torch.nn as nn
from django.db import connection
import datetime
from .models import Coordinate, Upload, Filename
from django.core.cache import cache
from django.db.models import Min
from django.http import JsonResponse
from openpyxl import load_workbook, Workbook
import pandas as pd
import torch
from django.conf import settings
import os
import joblib
from django.db.models import Sum
from datetime import datetime
from django.utils import timezone
from django.db.models import Count

connect = pymysql.connect(host='localhost', port=3306, user='root', password='root', db='track', charset='utf8')

cursor = connect.cursor()
data_df = None
page_data = []
front_table = None

def index(request):
    return render(request, 'index.html')

def diabetes(request):
    return render(request, 'Diabetes.html')

def navigation(request):
    return render(request, 'navigation.html')

def map(request):
    return get_uploads_with_times(request)

def about(request):
    return render(request, 'about.html')

def teaching(request):
    return render(request, 'teaching.html')



# 定义R^2损失函数
class R2Loss(torch.nn.Module):
    def __init__(self):
        super(R2Loss, self).__init__()

    def forward(self, y_pred, y_true):
        y_mean = torch.mean(y_true)
        SS_tot = torch.sum((y_true - y_mean) ** 2)
        SS_res = torch.sum((y_true - y_pred) ** 2)
        r2_score = 1 - SS_res / (SS_tot + 1e-8)  # 添加1e-8以避免除零错误
        return -r2_score
# 定义神经网络模型
# 定义MLP模型
# 定义注意力机制块
class AttentionBlock(nn.Module):
    def __init__(self, in_features):
        super(AttentionBlock, self).__init__()
        self.attn = nn.Linear(in_features, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        # 假设x的维度是 [batch_size, in_features]
        attention = self.attn(x)
        attention = self.sigmoid(attention)
        # 将注意力权重应用到输入x上
#         return x * attention.unsqueeze(2)
        return x * attention

# 定义残差块
class ResidualBlock(nn.Module):
    def __init__(self, in_features, out_features):
        super(ResidualBlock, self).__init__()
        self.match_dim = nn.Linear(in_features, out_features)
        self.fc1 = nn.Linear(in_features, out_features)
        self.relu = nn.ReLU()
        self.fc2 = nn.Linear(out_features, out_features)
        self.attention = AttentionBlock(out_features)

    def forward(self, x):
#         residual = x
        residual = self.match_dim(x)
        out = self.relu(self.fc1(x))
        out = self.fc2(out)
        out = self.attention(out)
        out = out + residual  # 残差连接
        return out

class FuelConsumptionModel(nn.Module):
    def __init__(self):
        super(FuelConsumptionModel, self).__init__()
        self.fc1 = nn.Linear(10, 16)  # 第一个全连接层，从10个特征到16个
        self.residual1 = ResidualBlock(16, 32)  # 扩展到32个特征
        self.fc2 = nn.Linear(32, 64)  # 进一步扩展到64个特征
        self.residual2 = ResidualBlock(64, 64)  # 保持64个特征
        self.fc3 = nn.Linear(64, 32)  # 减少回32个特征
        self.residual3 = ResidualBlock(32, 16)  
        self.fc4 = nn.Linear(16, 1)  # 最后的全连接层输出一个预测值

    def forward(self, x):
        # 假设x的维度是 [batch_size, 10]
        out = torch.relu(self.fc1(x))  # 通过第一个全连接层
        out = self.residual1(out)  # 通过第一个残差块
        out = torch.relu(self.fc2(out))  # 通过第二个全连接层
        out = self.residual2(out)  # 通过第二个残差块
        out = torch.relu(self.fc3(out))  # 通过第三个全连接层
        out = self.residual3(out)  # 通过第三个残差块
        out = self.fc4(out)  # 最后的全连接层输出预测值
        return out

@csrf_exempt  # 使用 csrf_exempt 装饰器避免 CSRF 验证（仅在开发过程中使用）
def predict(request):
    if request.method == 'POST':
        average_speed = request.POST.get('input1')
        speed_std = request.POST.get('input2')
        average_acceleration = request.POST.get('input3')
        rotation_speed = request.POST.get('input4')
        torque = request.POST.get('input5')
        acceleration_ratio = request.POST.get('input6')
        deceleration_ratio = request.POST.get('input7')
        constant_speed_ratio = request.POST.get('input8')
        idle_speed_ratio = request.POST.get('input9')
        power = request.POST.get('input10')
        
        # 确保所有数据都已收到
        if all([average_speed, speed_std, average_acceleration, rotation_speed, torque,
                acceleration_ratio, deceleration_ratio, constant_speed_ratio,
                idle_speed_ratio, power]):
            # 将数据转换为浮点数
            input_data = [float(average_speed), float(speed_std),
                          float(average_acceleration), float(rotation_speed),
                          float(torque), float(acceleration_ratio),
                          float(deceleration_ratio), float(constant_speed_ratio),
                          float(idle_speed_ratio), float(power)]
            print(input_data)
            # 创建模型实例
            model = FuelConsumptionModel()
            try:
                model.load_state_dict(torch.load('home/best_model.pth')) #最好模型参数是在kaggle上用gpu跑的，自己电脑上无gpu，这里无法载入，best_model.pth不是最好模型参数
                #若要使用最好模型，载入‘home/best_model1.pth’
            except Exception as e:
                print("Error loading model:", e)
                return JsonResponse({'error': 'Model loading failed'}, status=500)
            model.eval()  # 设置为评估模式

            scaler = joblib.load('home/scaler.joblib')
            input_array = scaler.transform([np.array(input_data)]).astype(np.float32)
            print(input_array)
            input_array = torch.from_numpy(input_array)  # 转换为torch张量
            # 准备输入数据
            # input_array = torch.tensor(input_data).type(torch.float32)  # 转换为张量和浮点型

            # 使用模型进行预测
            with torch.no_grad():
                prediction = model(input_array)

            # 获取预测结果
            result = prediction.item()

            # 将结果转换为 JSON 响应
            response_data = {'prediction': result}

            # 返回 JSON 响应
            return JsonResponse(response_data, safe=False)
        else:
            # 如果数据不完整，返回错误响应
            return JsonResponse({'error': 'Invalid data received'}, status=400)
    else:
        # 如果不是 POST 请求，返回方法不被允许的错误
        return JsonResponse({'error': 'Method Not Allowed'}, status=405)

@csrf_exempt
def predict1(request):
    if request.method == 'POST':
        fileobject = request.FILES.get("exc")
        if not fileobject:
            return JsonResponse({'error': '未上传文件'}, status=404)

        filename = fileobject.name
        wb = load_workbook(fileobject)
        sheet = wb.active
        model = FuelConsumptionModel()  # 假设这是你的模型类

        # 加载模型参数
        model.load_state_dict(torch.load('home/best_model.pth'))
        model.eval()  # 设置为评估模式
        scaler = joblib.load('home/scaler.joblib')
            
        # 确定新列的列号（在现有最大列号之后）
        last_column = sheet.max_column + 1

        # 遍历工作表中的每一行（跳过第一行标题）
        for row in sheet.iter_rows(min_row=2):
            # 从单元格中提取数据并转换为浮点数
            input_data = [float(cell.value) for cell in row]
            if len(input_data) == 10:
                # 将10列数据转换为张量
                input_array = scaler.transform([np.array(input_data)]).astype(np.float32)
                input_array = torch.from_numpy(input_array)  # 转换为torch张量
                
                # 使用模型进行预测
                with torch.no_grad():
                    prediction = model(input_array)
                
                # 获取预测结果
                prediction_value = prediction.item()

                # 将预测结果写入新的列
                sheet.cell(row=row[0].row, column=last_column).value = prediction_value

        # 保存工作簿到新文件
        file_path = os.path.join(settings.MEDIA_ROOT, 'updated_' + filename)
        wb.save(file_path)
        
        # 生成下载URL
        download_url = os.path.join(settings.MEDIA_URL, 'updated_' + filename)
        return JsonResponse({'download_url': download_url}, status=200)
    else:
        return JsonResponse({'error': 'Method Not Allowed'}, status=405)

@csrf_exempt
def predict_row(model, row):
    # 将行数据转换为模型输入
    input_data = [float(row[i]) for i in range(len(row)-1)]  # 假设最后一列是预测结果列
    input_array = torch.tensor(input_data).type(torch.float32)
    
    # 使用模型进行预测
    with torch.no_grad():
        prediction = model(input_array)
        return prediction.item()

@csrf_exempt
def ct(request):
    if request.method == 'POST':
        fileobject = request.FILES.get("exc")
        if not fileobject:
            return JsonResponse({'error': '未上传文件'}, status=404)
        filename = fileobject.name
        wb = load_workbook(fileobject)
        sheet = wb.worksheets[0]
        file = models.Filename.objects.create(
            name=filename,  
        )
        file.save()
        models.Coordinate.objects.all().delete()
        for row in sheet.iter_rows(min_row=2):
            text_lng = row[0].value
            text_lat = row[1].value
            text_time = row[2].value
            text_speed = row[3].value
            text_direction = row[4].value
            text_height = row[5].value
            text_did = row[6].value
            text_flag = row[7].value
            data = CoordinateForm(data={'lng':text_lng,'lat':text_lat,'time':text_time,'speed':text_speed,'direction':text_direction,'height':text_height,'did':text_did,'flag':text_flag})
            data.save()
        return redirect('/map/')
    else:
        # 如果请求不是 POST 请求，可以返回错误信息或处理逻辑
        return HttpResponseNotAllowed(['POST'])
    
@csrf_exempt
def coor(request):
    if request.method == 'POST':
        # 从数据库获取当前用户的上传次数
        max_upload_count = models.Upload.objects.all().aggregate(Max('upload_count'))['upload_count__max'] or 0
        upload_count = max_upload_count + 1

        # 获取所有经纬度数据
        locations = models.Coordinate.objects.all().values_list('lng', 'lat')
        time = models.Coordinate.objects.all().values_list('time')
        speed = models.Coordinate.objects.all().values_list('speed')
        direction = models.Coordinate.objects.all().values_list('direction')
        height = models.Coordinate.objects.all().values_list('height')
        did = models.Coordinate.objects.all().values_list('did')
        flag = models.Coordinate.objects.all().values_list('flag')
        name = models.Filename.objects.first().name

        # 假设您的char字段名为time_char，并且存储的时间字符串格式是一致的
        grouped_coordinates = models.Coordinate.objects \
            .values('did', 'time') \
            .annotate(first_time_char=Min('time')) \
            .order_by('did')

        first_coordinates_list = []

        for group in grouped_coordinates:
            # 使用annotate的first_time_char作为筛选条件，获取每个分组的第一行数据
            first_coordinate = models.Coordinate.objects.filter(
                did=group['did'],
                time=group['first_time_char']
            ).values_list('lng', 'lat').first()

            if first_coordinate:
                #first_coordinates_list.append(list(first_coordinate))
                first_coordinates_list.append({'did': group['did'], 'coordinate': list(first_coordinate)})
        
        # 将查询集转换为列表的列表，每个内部列表包含一个经纬度对
        coordinates_list = [list(coord) for coord in locations]
        time_list = [list(data) for data in time]
        speed_list = [list(data) for data in speed]
        direction_list = [list(data) for data in direction]
        height_list = [list(data) for data in height]
        did_list = [list(data) for data in did]
        flag_list = [list(data) for data in flag]
       
        # 创建一个新的 Upload 实例
        # 假设 Upload 模型有一个 JSONField 叫 coordinates_field
        upload = models.Upload.objects.create(
            coordinates=json.dumps(coordinates_list),  # 保存 JSON 字符串
            upload_count=upload_count,
            time=json.dumps(time_list), 
            speed=json.dumps(speed_list), 
            direction=json.dumps(direction_list), 
            height=json.dumps(height_list), 
            did=json.dumps(did_list), 
            flag=json.dumps(flag_list), 
            file_name=name,
            start_coor=json.dumps(first_coordinates_list),
        )
        models.Coordinate.objects.all().delete()
        models.Filename.objects.all().delete()

        response_data = {
            'upload': {
                'coordinates': upload.coordinates,
                'upload_count': upload.upload_count,
                'time': upload.time,
                'speed': upload.speed,
                'direction': upload.direction,
                'height': upload.height,
                'did': upload.did,
                'flag': upload.flag,
                'start_coor': upload.start_coor,
                'file_name': upload.file_name,
                'created_at': upload.created_at,
            }
        }
        #print(response_data)  # 确保这里没有 QuerySet 对象       
        # 返回 JSON 响应
        return JsonResponse(response_data)
    else:
        # 如果不是 POST 请求，返回 405 方法不被允许 响应
        return HttpResponseNotAllowed(['POST'])

@csrf_exempt
def history(request):
    return render(request, 'history.html')

@csrf_exempt
def show_data_info(request):
    # 使用 ORM 获取查询结果
    home_uploads = Upload.objects.all().values_list('file_name', 'created_at', 'start_coor','upload_count')
    # values_list() 返回一个QuerySet，其中每个元素都是一个包含指定字段值的元组
    
    # 将查询结果转换为列表
    res1 = list(home_uploads)

    # 遍历列表索引并修改元组
    for index, item in enumerate(res1):
        if item[2]:  # 假设 item[2] 是 start_coor 字段
            # 解析 JSON 字符串为 Python 对象
            res1[index] = list(item)  # 创建一个新的列表以便修改
            res1[index][2] = json.loads(item[2])  # 更新 start_coor 字段
    context = {
        'data': res1
    }
    return render(request, 'history.html', context)

@csrf_exempt
def delete_data(request):
    if request.method == 'POST':
        count = request.POST.get('data_id')
        # 使用filter来检索所有匹配data_id的记录，然后删除它们
        Upload.objects.filter(upload_count=count).delete()
        return redirect('/history/')  # 根据实际的URL名称进行重定向

@csrf_exempt
def detail_data(request):
    if request.method == 'POST':
        count = request.POST['data_id']
        print(count)
        home_upload = Upload.objects.get(upload_count=count)
        if isinstance(home_upload.coordinates, str):
            try:
                home_upload.coordinates = json.loads(home_upload.coordinates)
            except json.JSONDecodeError:
                # 如果 JSON 解码失败，返回错误信息或进行其他处理
                return JsonResponse({'error': 'Invalid JSON data for coordinates'}, status=400)
    
        #确保 coordinates 是一个列表的列表
        if not all(isinstance(coord, list) and len(coord) == 2 for coord in home_upload.coordinates):
            # 如果数据结构不正确，返回错误信息或进行其他处理
            return JsonResponse({'error': 'Invalid coordinates data structure'}, status=400)
        
        if isinstance(home_upload.time, str):
            try:
                home_upload.time = json.loads(home_upload.time)
            except json.JSONDecodeError:
                # 如果 JSON 解码失败，返回错误信息或进行其他处理
                return JsonResponse({'error': 'Invalid JSON data for time'}, status=400)
         
        if isinstance(home_upload.speed, str):
            try:
                home_upload.speed = json.loads(home_upload.speed)
            except json.JSONDecodeError:
                # 如果 JSON 解码失败，返回错误信息或进行其他处理
                return JsonResponse({'error': 'Invalid JSON data for speed'}, status=400)

        if isinstance(home_upload.direction, str):
            try:
                home_upload.direction = json.loads(home_upload.direction)
            except json.JSONDecodeError:
                # 如果 JSON 解码失败，返回错误信息或进行其他处理
                return JsonResponse({'error': 'Invalid JSON data for direction'}, status=400)
    
        if isinstance(home_upload.height, str):
            try:
                home_upload.height = json.loads(home_upload.height)
            except json.JSONDecodeError:
                # 如果 JSON 解码失败，返回错误信息或进行其他处理
                return JsonResponse({'error': 'Invalid JSON data for height'}, status=400)

        if isinstance(home_upload.did, str):
            try:
                home_upload.did = json.loads(home_upload.did)
            except json.JSONDecodeError:
                # 如果 JSON 解码失败，返回错误信息或进行其他处理
                return JsonResponse({'error': 'Invalid JSON data for did'}, status=400)

        if isinstance(home_upload.flag, str):
            try:
                home_upload.flag = json.loads(home_upload.flag)
            except json.JSONDecodeError:
                # 如果 JSON 解码失败，返回错误信息或进行其他处理
                return JsonResponse({'error': 'Invalid JSON data for flag'}, status=400)
    
        context = {
            'data1': home_upload,
        }
        return render(request, 'detail.html', context)
    
@csrf_exempt
def get_uploads_with_times(request):
    # 获取所有上传记录的ID和创建时间
    uploads = Upload.objects.all().values('file_name', 'created_at')
    uploads_list = list(uploads)  # 将查询集转换为列表

    # 将查询结果格式化为JSON
    uploads_json = [{
        'file_name': upload['file_name'],
        'created_at': upload['created_at'].strftime('%Y-%m-%d %H:%M:%S')  # 格式化日期时间
    } for upload in uploads_list]

    # 返回JSON响应
    return render(request, 'map.html', {"uploads": uploads_json})

@csrf_exempt
def get_map_data(request):
    try:
        if request.method == 'POST' and request.body:
            data = json.loads(request.body)
            selected_ids = data.get('selected_uploads', [])
            selected_uploads = Upload.objects.filter(file_name__in=selected_ids)
            # 准备要返回的数据，这里需要根据您的具体需求来构造数据格式
            map_data_list = []
            for upload in selected_uploads:
                map_data = {
                    'created_at': upload.created_at,
                    'upload_id': upload.id,
                    'coordinates': upload.coordinates,  # 确保这是您想要的格式
                    'upload_count': upload.upload_count,
                    'time': upload.time,
                    'speed': upload.speed,
                    'direction': upload.direction,
                    'height': upload.height,
                    'did': upload.did,
                    'flag': upload.flag,
                    'file_name': upload.file_name,
                    'start_coor': upload.start_coor,
                }
                map_data_list.append(map_data)
            return JsonResponse({'uploads': map_data_list}, safe=False)  # safe=False因为返回的是数组
    except Upload.DoesNotExist:
        return JsonResponse({'error': 'Upload data not found.'}, status=404)

@csrf_exempt
def get_daily_upload_counts(request):
        # 假设前端传递了start_date和end_date作为查询参数
    # start_date_str = request.GET.get('start_date', default_start_date)  # 默认_start_date为某个默认值或None
    # end_date_str = request.GET.get('end_date', default_end_date)      # 默认_end_date为某个默认值或None
    start_date_str = request.GET.get('start_date')  # 默认_start_date为某个默认值或None
    end_date_str = request.GET.get('end_date')      # 默认_end_date为某个默认值或None
    # 将字符串日期转换为datetime对象
    start_date = timezone.make_aware(datetime.strptime(start_date_str, '%Y-%m-%d'))
    end_date = timezone.make_aware(datetime.strptime(end_date_str, '%Y-%m-%d'))
    data1 = Upload.objects.filter(created_at__range=(start_date, end_date)).values('created_at__date')
    print(data1)
    data2 = Upload.objects.filter(created_at__range=(start_date, end_date))
    print(data2)
    # 获取每天的上传总数
    daily_counts = Upload.objects.filter(created_at__range=(start_date, end_date)).values('created_at__date').annotate(total_uploads=Count('id'))  # 使用Count来计算每天的记录数)  # 计算每天的上传总数
    print(daily_counts)
    # 序列化为JSON格式
    data = [{
        'created_at__date': str(item['created_at__date']),  # 将日期转换为字符串
        'total_uploads': item['total_uploads']
    } for item in daily_counts]
    print(data)
    # 返回JSONResponse
    return JsonResponse(data, safe=False)